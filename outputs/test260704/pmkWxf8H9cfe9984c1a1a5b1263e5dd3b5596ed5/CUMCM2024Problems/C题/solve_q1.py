#!/usr/bin/env python3
"""
CUMCM 2024 Problem C - Question 1: Optimal Crop Planting Strategy
Maximize revenue for 2024-2030 under two scenarios:
  (1) Excess production wasted
  (2) Excess production sold at 50% price

Method: Mixed Integer Linear Programming (PuLP + CBC)
"""

import openpyxl
import json
import os
import sys
from collections import defaultdict
from itertools import product

# Suppress PuLP logging
import logging
logging.basicConfig(level=logging.WARNING)

try:
    from pulp import *
except ImportError:
    os.system("pip install pulp -q")
    from pulp import *

# ============================================================
# SECTION 1: DATA PARSING
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = BASE_DIR  # same directory

def parse_plots(filepath):
    """Parse 附件1 sheet 1: plot info"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    plots = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        name = str(row[0]).strip()
        ptype = str(row[1]).strip()
        area = float(row[2])
        plots.append({'name': name, 'type': ptype, 'area': area})
    return plots

def parse_crops(filepath):
    """Parse 附件1 sheet 2: crop info"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[1]]
    crops = []
    current_plantable = ""
    current_note = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        cid = int(row[0])
        cname = str(row[1]).strip()
        ctype = str(row[2]).strip()
        if row[3] and str(row[3]).strip():
            current_plantable = str(row[3]).strip()
        if row[4] and str(row[4]).strip():
            current_note = str(row[4]).strip()
        crops.append({
            'id': cid, 'name': cname, 'type': ctype,
            'plantable': current_plantable,
            'note': current_note
        })
    return crops

def parse_planting_2023(filepath):
    """Parse 附件2 sheet 1: 2023 planting data"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[1] is None:
            break
        records.append({
            'plot': str(row[0]).strip() if row[0] else None,
            'crop_id': int(row[1]) if row[1] else None,
            'crop_name': str(row[2]).strip() if row[2] else None,
            'crop_type': str(row[3]).strip() if row[3] else None,
            'area': float(row[4]) if row[4] else 0,
            'season': str(row[5]).strip() if row[5] else None
        })
    return records

def parse_stats(filepath):
    """Parse 附件2 sheet 2: yield/cost/price statistics"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        price_str = str(row[7]).strip() if row[7] else ''
        # Parse price range "min-max" -> midpoint
        if '-' in price_str:
            parts = price_str.split('-')
            try:
                price_mid = (float(parts[0]) + float(parts[1])) / 2
            except:
                price_mid = 0
        else:
            try:
                price_mid = float(price_str)
            except:
                price_mid = 0

        records.append({
            'seq': int(row[0]),
            'crop_id': int(row[1]),
            'crop_name': str(row[2]).strip(),
            'land_type': str(row[3]).strip(),
            'season': str(row[4]).strip(),
            'yield_per_mu': float(row[5]),
            'cost_per_mu': float(row[6]),
            'price_mid': price_mid
        })
    return records

# Parse all data
plots = parse_plots(os.path.join(DATA_DIR, '附件1.xlsx'))
crops = parse_crops(os.path.join(DATA_DIR, '附件1.xlsx'))
planting_2023 = parse_planting_2023(os.path.join(DATA_DIR, '附件2.xlsx'))
stats = parse_stats(os.path.join(DATA_DIR, '附件2.xlsx'))

print(f"Plots: {len(plots)}, Crops: {len(crops)}, 2023 records: {len(planting_2023)}, Stat records: {len(stats)}")

# ============================================================
# SECTION 2: DATA STRUCTURES
# ============================================================

# Map land type to standardized name
LAND_TYPE_MAP = {
    '平旱地': 'pinghan', '梯田': 'titian', '山坡地': 'shanpo',
    '水浇地': 'shuijiao', '普通大棚': 'putongdp', '智慧大棚': 'zhihuidp'
}

# Standardize plot types
for p in plots:
    raw = p['type'].strip()
    p['type_std'] = LAND_TYPE_MAP.get(raw, raw)

# Crop categories
LEGUME_IDS = {1, 2, 3, 4, 5, 17, 18, 19}  # 豆类: 黄豆,黑豆,红豆,绿豆,芸豆,豇豆,刀豆,芸豆(蔬菜豆类)
# Note: 17(豇豆),18(刀豆),19(芸豆) are 蔬菜(豆类) - vegetable beans
# They can be planted on 水浇地/大棚 and count toward legume rotation requirement
GRAIN_IDS = set(range(6, 16))  # 粮食: 6-15
RICE_ID = 16  # 水稻
VEG_S1_IDS = set(range(17, 35))  # 第一季蔬菜 (含水浇地第二季限定品种... no, those are separate)
VEG_S2_ONLY_IDS = {35, 36, 37}  # 大白菜,白萝卜,红萝卜 - 水浇地第二季only
FUNGI_IDS = {38, 39, 40, 41}  # 食用菌

# Build stat lookup: (crop_id, land_type_std, season) -> {yield, cost, price}
stat_lookup = {}
for s in stats:
    lt = s['land_type'].strip()
    lt_std = LAND_TYPE_MAP.get(lt, lt)
    key = (s['crop_id'], lt_std, s['season'])
    stat_lookup[key] = {
        'yield_per_mu': s['yield_per_mu'],
        'cost_per_mu': s['cost_per_mu'],
        'price': s['price_mid']
    }

# Add 智慧大棚 first season = same as 普通大棚 first season
for cid in VEG_S1_IDS:
    key_src = (cid, 'putongdp', '第一季')
    key_dst = (cid, 'zhihuidp', '第一季')
    if key_src in stat_lookup and key_dst not in stat_lookup:
        stat_lookup[key_dst] = stat_lookup[key_src].copy()

# Build compatibility matrix: which (crop_id, season) can be planted on each plot type
# Rules:
# - 平旱地/梯田/山坡地: crops 1-15, 单季 only
# - 水浇地: single season rice (16), OR two-season vegetables (17-34 first, 35-37 second)
# - 普通大棚: first season vegetables (17-34), second season edible fungi (38-41)
# - 智慧大棚: both seasons vegetables (17-34)

def get_allowed_crops_seasons(plot_type_std):
    """Return list of (crop_id, season) tuples allowed on this plot type"""
    allowed = []
    if plot_type_std in ('pinghan', 'titian', 'shanpo'):
        # Single season, grain+legume crops 1-15
        for cid in range(1, 16):
            allowed.append((cid, '单季'))
    elif plot_type_std == 'shuijiao':
        # Rice single season
        allowed.append((16, '单季'))
        # Vegetables first season
        for cid in range(17, 35):
            allowed.append((cid, '第一季'))
        # Vegetables second season (only 35-37)
        for cid in (35, 36, 37):
            allowed.append((cid, '第二季'))
    elif plot_type_std == 'putongdp':
        # Vegetables first season
        for cid in range(17, 35):
            allowed.append((cid, '第一季'))
        # Edible fungi second season
        for cid in (38, 39, 40, 41):
            allowed.append((cid, '第二季'))
    elif plot_type_std == 'zhihuidp':
        # Vegetables both seasons
        for cid in range(17, 35):
            allowed.append((cid, '第一季'))
            allowed.append((cid, '第二季'))
    return allowed

# ============================================================
# SECTION 3: COMPUTE EXPECTED SALES FROM 2023
# ============================================================

# Reconcile plot type standardization for stat lookup
plot_type_map = {p['name']: p['type_std'] for p in plots}
# print("Plot type map:", plot_type_map)

# Compute 2023 total production per crop
prod_2023 = defaultdict(float)
for rec in planting_2023:
    if rec['crop_id'] is None:
        continue
    cid = rec['crop_id']
    pname = rec['plot']
    season = rec['season']
    area = rec['area']
    pt = plot_type_map.get(pname, 'unknown')
    key = (cid, pt, season)
    if key in stat_lookup:
        yld = stat_lookup[key]['yield_per_mu']
        prod_2023[cid] += yld * area

expected_sales = dict(prod_2023)
print("\nExpected sales (2023 production, jin):")
for cid in sorted(expected_sales.keys()):
    cname = next((c['name'] for c in crops if c['id'] == cid), f"crop_{cid}")
    print(f"  Crop {cid} ({cname}): {expected_sales[cid]:.0f} jin")

# ============================================================
# SECTION 4: BUILD OPTIMIZATION MODEL
# ============================================================

YEARS = list(range(2024, 2031))  # 2024-2030
CROP_IDS_ALL = set(range(1, 42))

def build_and_solve(scenario=1, verbose=True):
    """
    scenario 1: excess wasted
    scenario 2: excess sold at 50%
    """

    # Create LP problem
    prob = LpProblem(f"CUMCM2024C_Q1_S{scenario}", LpMaximize)

    # ---- Decision Variables ----
    # x[pname, year, cid, season] = planting area
    x = {}
    for p in plots:
        pn = p['name']
        pt = p['type_std']
        allowed = get_allowed_crops_seasons(pt)
        for yr in YEARS:
            for (cid, season) in allowed:
                # Check if stat data exists
                key = (cid, pt, season)
                if key in stat_lookup:
                    x[pn, yr, cid, season] = LpVariable(
                        f"x_{pn}_{yr}_{cid}_{season}", lowBound=0, upBound=p['area']
                    )

    # NOTE: No binary variables needed. Rotation enforced directly via
    # continuous constraints on total legume area per 3-year window.

    # Production tracking per crop per year
    prod = {}
    for yr in YEARS:
        for cid in CROP_IDS_ALL:
            prod[yr, cid] = LpVariable(f"prod_{yr}_{cid}", lowBound=0)

    # Sales tracking (scenario 1) or excess tracking (scenario 2)
    if scenario == 1:
        sold = {}
        for yr in YEARS:
            for cid in CROP_IDS_ALL:
                if cid in expected_sales:
                    sold[yr, cid] = LpVariable(f"sold_{yr}_{cid}", lowBound=0,
                                                upBound=expected_sales[cid])
    else:  # scenario 2
        excess = {}
        for yr in YEARS:
            for cid in CROP_IDS_ALL:
                if cid in expected_sales:
                    excess[yr, cid] = LpVariable(f"excess_{yr}_{cid}", lowBound=0)

    # ---- Constraints ----

    # (C1) Plot area constraints
    for p in plots:
        pn = p['name']
        for yr in YEARS:
            # Sum of all crop areas on this plot in this year ≤ plot area
            relevant_vars = []
            for (pn2, yr2, cid, season), var in x.items():
                if pn2 == pn and yr2 == yr:
                    relevant_vars.append(var)
            if relevant_vars:
                prob += lpSum(relevant_vars) <= p['area'], f"area_{pn}_{yr}"

    # (C2) Water-irrigated land: second season ≤ first season vegetable area
    for p in plots:
        if p['type_std'] != 'shuijiao':
            continue
        pn = p['name']
        for yr in YEARS:
            s1_vars = [x.get((pn, yr, cid, '第一季'), None) for cid in range(17, 35)]
            s2_vars = [x.get((pn, yr, cid, '第二季'), None) for cid in (35, 36, 37)]
            s1_vars = [v for v in s1_vars if v is not None]
            s2_vars = [v for v in s2_vars if v is not None]
            if s1_vars and s2_vars:
                prob += lpSum(s2_vars) <= lpSum(s1_vars), f"shuijiao_s2_{pn}_{yr}"

    # (C3) Production definition
    for yr in YEARS:
        for cid in CROP_IDS_ALL:
            prod_terms = []
            for (pn2, yr2, cid2, season), var in x.items():
                if yr2 == yr and cid2 == cid:
                    ptype = plot_type_map[pn2]
                    key = (cid, ptype, season)
                    if key in stat_lookup:
                        yld = stat_lookup[key]['yield_per_mu']
                        prod_terms.append(yld * var)
            if prod_terms:
                prob += prod[yr, cid] == lpSum(prod_terms), f"prod_def_{yr}_{cid}"
            else:
                prob += prod[yr, cid] == 0, f"prod_zero_{yr}_{cid}"

    # (C4) Sales/excess constraints
    if scenario == 1:
        for yr in YEARS:
            for cid in expected_sales:
                if (yr, cid) in prod and (yr, cid) in sold:
                    prob += sold[yr, cid] <= prod[yr, cid], f"sold_le_prod_{yr}_{cid}"
    else:  # scenario 2
        for yr in YEARS:
            for cid in expected_sales:
                if (yr, cid) in prod and (yr, cid) in excess:
                    prob += excess[yr, cid] >= prod[yr, cid] - expected_sales[cid], \
                        f"excess_ge_{yr}_{cid}"

    # (C5) Rotation constraint: every plot must have legumes at least once per 3-year window
    # Direct constraint on total legume area (no binary variables needed)
    # "每个地块（含大棚）在三年内至少种植一次豆类作物"

    MIN_LEGUME_AREA = 0.1  # minimum mu of legumes per 3-year window

    # Track which plots had legumes in 2023 (from historical data)
    legume_2023 = set()
    legume_area_2023 = defaultdict(float)
    for rec in planting_2023:
        if rec['crop_id'] in LEGUME_IDS:
            legume_2023.add(rec['plot'])
            legume_area_2023[rec['plot']] += rec['area']

    for p in plots:
        pn = p['name']
        pt = p['type_std']

        # Collect all legume decision variables for this plot across all years
        legume_by_year = {}
        for yr in YEARS:
            yr_legume_vars = []
            for cid in LEGUME_IDS:
                for season in ['单季', '第一季', '第二季']:
                    v = x.get((pn, yr, cid, season), None)
                    if v is not None:
                        yr_legume_vars.append(v)
            if yr_legume_vars:
                legume_by_year[yr] = lpSum(yr_legume_vars)

        # 3-year window constraints
        # Window 1: [2023, 2024, 2025]
        # If legume_area in 2023 >= MIN_LEGUME_AREA, constraint already satisfied
        # Otherwise, need enough in 2024+2025
        had_2023 = legume_area_2023.get(pn, 0)
        if had_2023 >= MIN_LEGUME_AREA:
            # 2023 already satisfies [2023,2024,2025] window
            pass
        else:
            # Need legume in 2024 or 2025
            deficit = MIN_LEGUME_AREA - had_2023
            yr_vars = []
            for yr in [2024, 2025]:
                if yr in legume_by_year:
                    yr_vars.append(legume_by_year[yr])
            if yr_vars:
                prob += lpSum(yr_vars) >= deficit, f"rot_2023_{pn}"

        # Windows [2024,2025,2026] through [2028,2029,2030]
        for t_start in range(2024, 2029):
            yr_vars = []
            for yr in [t_start, t_start+1, t_start+2]:
                if yr in legume_by_year:
                    yr_vars.append(legume_by_year[yr])
            if yr_vars:
                prob += lpSum(yr_vars) >= MIN_LEGUME_AREA, f"rot_{t_start}_{pn}"

    # ---- Objective Function ----
    obj_terms = []

    for yr in YEARS:
        # Revenue from sales
        if scenario == 1:
            for cid in expected_sales:
                if (yr, cid) in sold:
                    # Get price (use one representative price from stat data)
                    price = None
                    for (c, lt, s), info in stat_lookup.items():
                        if c == cid:
                            price = info['price']
                            break
                    if price is None:
                        price = 0
                    obj_terms.append(price * sold[yr, cid])
        else:  # scenario 2
            for cid in expected_sales:
                if (yr, cid) in prod:
                    price = None
                    for (c, lt, s), info in stat_lookup.items():
                        if c == cid:
                            price = info['price']
                            break
                    if price is None:
                        price = 0
                    # Full price on all production minus half-price penalty on excess
                    obj_terms.append(price * prod[yr, cid])
                    if (yr, cid) in excess:
                        obj_terms.append(-0.5 * price * excess[yr, cid])

        # Subtract costs
        for (pn, yr2, cid, season), var in x.items():
            if yr2 == yr:
                ptype = plot_type_map[pn]
                key = (cid, ptype, season)
                if key in stat_lookup:
                    cost = stat_lookup[key]['cost_per_mu']
                    obj_terms.append(-cost * var)

    prob += lpSum(obj_terms)

    # ---- Solve ----
    if verbose:
        print(f"\nSolving scenario {scenario}...")
        print(f"  Variables: {len(prob.variables())}")
        print(f"  Constraints: {len(prob.constraints)}")

    solver = PULP_CBC_CMD(msg=verbose)
    prob.solve(solver)

    if verbose:
        print(f"  Status: {LpStatus[prob.status]}")
        print(f"  Objective: {value(prob.objective):,.0f} yuan")

    return prob, x, prod, sold if scenario == 1 else excess

# ============================================================
# SECTION 5: SOLVE BOTH SCENARIOS AND OUTPUT
# ============================================================

print("\n" + "="*60)
print("SCENARIO 1: Excess production wasted")
print("="*60)
prob1, x1, prod1, sold1 = build_and_solve(scenario=1, verbose=True)

print("\n" + "="*60)
print("SCENARIO 2: Excess sold at 50% price")
print("="*60)
prob2, x2, prod2, excess2 = build_and_solve(scenario=2, verbose=True)

# ============================================================
# SECTION 6: EXTRACT RESULTS AND SAVE TO EXCEL
# ============================================================

def extract_solution(x_vars, prod_vars, aux_vars, scenario, prob):
    """Extract planting decisions for output"""
    # results[year][plot_name] = {crop_id: area} for single season
    # For multi-season plots, store separately

    result = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # result[year][plot][crop_id] = total area (for single-season plots)
    # For multi-season, we need to distinguish seasons

    result_detail = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
    # result_detail[year][plot][season][crop_id] = area

    for (pn, yr, cid, season), var in x_vars.items():
        val = value(var)
        if val is not None and val > 0.001:  # threshold for numerical noise
            result_detail[yr][pn][season][cid] = val
            result[yr][pn][cid] += val

    return result, result_detail

def write_result_excel(result_detail, output_path, template_path=None):
    """Write results to Excel in the format of result1_1/result1_2/result2"""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Crop ID to name mapping
    crop_id_to_name = {c['id']: c['name'] for c in crops}

    # All crop IDs in order
    all_crop_ids = list(range(1, 42))

    # Plot names in order
    plot_names = [p['name'] for p in plots]
    plot_areas = {p['name']: p['area'] for p in plots}

    for yr in YEARS:
        ws = wb.create_sheet(title=str(yr))

        # Header row
        headers = ['', '地块名称'] + [crop_id_to_name.get(cid, str(cid)) for cid in all_crop_ids]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        # Data rows
        row = 2
        # Group plots by type for better readability
        season_label = '第一季'
        for pname in plot_names:
            ptype = plot_type_map.get(pname, '')

            if ptype in ('shuijiao', 'putongdp', 'zhihuidp'):
                # Multi-season plots: need separate rows for each season
                seasons = []
                if ptype == 'shuijiao':
                    # Check if there's rice (单季) or vegetables (第一季/第二季)
                    has_s1 = any(result_detail[yr][pname].get('第一季', {}).values())
                    has_s2 = any(result_detail[yr][pname].get('第二季', {}).values())
                    has_dj = any(result_detail[yr][pname].get('单季', {}).values())

                    if has_dj:
                        # Rice single season
                        ws.cell(row=row, column=1, value='第一季')
                        ws.cell(row=row, column=2, value=pname)
                        for cid in all_crop_ids:
                            area = result_detail[yr][pname].get('单季', {}).get(cid, 0)
                            if area > 0.001:
                                ws.cell(row=row, column=2+cid, value=area)
                        row += 1

                    if has_s1:
                        ws.cell(row=row, column=1, value='第一季')
                        ws.cell(row=row, column=2, value=pname)
                        for cid in all_crop_ids:
                            area = result_detail[yr][pname].get('第一季', {}).get(cid, 0)
                            if area > 0.001:
                                ws.cell(row=row, column=2+cid, value=area)
                        row += 1

                    if has_s2:
                        ws.cell(row=row, column=1, value='第二季')
                        ws.cell(row=row, column=2, value=pname)
                        for cid in all_crop_ids:
                            area = result_detail[yr][pname].get('第二季', {}).get(cid, 0)
                            if area > 0.001:
                                ws.cell(row=row, column=2+cid, value=area)
                        row += 1

                    if not has_dj and not has_s1 and not has_s2:
                        # Empty row
                        ws.cell(row=row, column=1, value='第一季')
                        ws.cell(row=row, column=2, value=pname)
                        row += 1

                elif ptype == 'putongdp':
                    has_s1 = any(result_detail[yr][pname].get('第一季', {}).values())
                    has_s2 = any(result_detail[yr][pname].get('第二季', {}).values())

                    if has_s1:
                        ws.cell(row=row, column=1, value='第一季')
                        ws.cell(row=row, column=2, value=pname)
                        for cid in all_crop_ids:
                            area = result_detail[yr][pname].get('第一季', {}).get(cid, 0)
                            if area > 0.001:
                                ws.cell(row=row, column=2+cid, value=area)
                        row += 1

                    if has_s2:
                        ws.cell(row=row, column=1, value='第二季')
                        ws.cell(row=row, column=2, value=pname)
                        for cid in all_crop_ids:
                            area = result_detail[yr][pname].get('第二季', {}).get(cid, 0)
                            if area > 0.001:
                                ws.cell(row=row, column=2+cid, value=area)
                        row += 1

                    if not has_s1 and not has_s2:
                        ws.cell(row=row, column=1, value='第一季')
                        ws.cell(row=row, column=2, value=pname)
                        row += 1

                elif ptype == 'zhihuidp':
                    has_s1 = any(result_detail[yr][pname].get('第一季', {}).values())
                    has_s2 = any(result_detail[yr][pname].get('第二季', {}).values())

                    if has_s1:
                        ws.cell(row=row, column=1, value='第一季')
                        ws.cell(row=row, column=2, value=pname)
                        for cid in all_crop_ids:
                            area = result_detail[yr][pname].get('第一季', {}).get(cid, 0)
                            if area > 0.001:
                                ws.cell(row=row, column=2+cid, value=area)
                        row += 1

                    if has_s2:
                        ws.cell(row=row, column=1, value='第二季')
                        ws.cell(row=row, column=2, value=pname)
                        for cid in all_crop_ids:
                            area = result_detail[yr][pname].get('第二季', {}).get(cid, 0)
                            if area > 0.001:
                                ws.cell(row=row, column=2+cid, value=area)
                        row += 1

                    if not has_s1 and not has_s2:
                        ws.cell(row=row, column=1, value='第一季')
                        ws.cell(row=row, column=2, value=pname)
                        row += 1
            else:
                # Single season plots
                ws.cell(row=row, column=1, value='单季')
                ws.cell(row=row, column=2, value=pname)
                for cid in all_crop_ids:
                    area = result_detail[yr][pname].get('单季', {}).get(cid, 0)
                    if area > 0.001:
                        ws.cell(row=row, column=2+cid, value=area)
                row += 1

    wb.save(output_path)
    return output_path

# Extract and save results
result1, detail1 = extract_solution(x1, prod1, sold1, 1, prob1)
result2, detail2 = extract_solution(x2, prod2, excess2, 2, prob2)

output_dir = os.path.join(DATA_DIR, 'C_ouput')
os.makedirs(output_dir, exist_ok=True)

out1 = write_result_excel(detail1, os.path.join(output_dir, 'result1_1.xlsx'))
out2 = write_result_excel(detail2, os.path.join(output_dir, 'result1_2.xlsx'))
print(f"\nResults saved to:")
print(f"  Scenario 1: {out1}")
print(f"  Scenario 2: {out2}")

# ============================================================
# SECTION 7: SUMMARY STATISTICS
# ============================================================

def compute_summary(result_detail, prob):
    """Compute yearly revenue and production summary"""
    summary = []
    for yr in YEARS:
        revenue = 0
        cost = 0
        production = defaultdict(float)

        for pn, seasons in result_detail[yr].items():
            ptype = plot_type_map[pn]
            for season, crops_dict in seasons.items():
                for cid, area in crops_dict.items():
                    key = (cid, ptype, season)
                    if key in stat_lookup:
                        yld = stat_lookup[key]['yield_per_mu']
                        cst = stat_lookup[key]['cost_per_mu']
                        prc = stat_lookup[key]['price']
                        prod_qty = yld * area
                        production[cid] += prod_qty
                        cost += cst * area

        # Compute revenue considering expected sales caps (scenario-dependent)
        rev = 0
        for cid, prod_qty in production.items():
            if cid in expected_sales:
                es = expected_sales[cid]
                prc = None
                for (c, lt, s), info in stat_lookup.items():
                    if c == cid:
                        prc = info['price']
                        break
                if prc is None:
                    prc = 0
                sold_qty = min(prod_qty, es)
                rev += sold_qty * prc

        summary.append({
            'year': yr,
            'revenue': rev,
            'cost': cost,
            'profit': rev - cost,
            'total_production': sum(production.values())
        })

    return summary

# Compute and print summaries
print("\n--- Scenario 1 Summary ---")
sum1 = compute_summary(detail1, prob1)
total_profit_1 = 0
for s in sum1:
    print(f"  {s['year']}: Revenue={s['revenue']:,.0f}, Cost={s['cost']:,.0f}, Profit={s['profit']:,.0f}")
    total_profit_1 += s['profit']
print(f"  TOTAL 2024-2030 Profit: {total_profit_1:,.0f} yuan")

print("\n--- Scenario 2 Summary ---")
sum2 = compute_summary(detail2, prob2)
total_profit_2 = 0
for s in sum2:
    print(f"  {s['year']}: Revenue={s['revenue']:,.0f}, Cost={s['cost']:,.0f}, Profit={s['profit']:,.0f}")
    total_profit_2 += s['profit']
print(f"  TOTAL 2024-2030 Profit: {total_profit_2:,.0f} yuan")

# Save summary JSON for report generation
summary_data = {
    'scenario1': {
        'yearly': [{'year': s['year'], 'revenue': round(s['revenue']),
                     'cost': round(s['cost']), 'profit': round(s['profit'])} for s in sum1],
        'total_profit': round(total_profit_1)
    },
    'scenario2': {
        'yearly': [{'year': s['year'], 'revenue': round(s['revenue']),
                     'cost': round(s['cost']), 'profit': round(s['profit'])} for s in sum2],
        'total_profit': round(total_profit_2)
    }
}

with open(os.path.join(output_dir, 'summary.json'), 'w', encoding='utf-8') as f:
    json.dump(summary_data, f, ensure_ascii=False, indent=2)

print(f"\nSummary saved to {os.path.join(output_dir, 'summary.json')}")
print("\nDone!")
